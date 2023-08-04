from datetime import datetime
import os
import openpyxl
from bs4 import BeautifulSoup
import requests

##! variable
# today = date.today().strftime("%Y-%m-%d")
today = datetime.now().strftime("%Y-%m-%d-%H:%M")
x = datetime.now().strftime("D= %Y-%m-%d-T= %H-%M-%S")
filename = f'Amazon_{x}.html'

url = 'https://www.amazon.com/s?k=ps5&rh=n%3A468642%2Cp_n_availability%3A2661601011%2Cp_36%3A2422979011&dc&ds=v1%3AWW9e45Z%2BxuwTcuTbWyWkW%2FcY7x0VQKZXRwcf0sq7%2FN4&crid=Y1KAPTIHUOJX&qid=1691163267&rnid=386453011&sprefix=ps%2Caps%2C498&ref=sr_nr_p_36_5'

headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36", "Accept-Encoding": "gzip, deflate",
           "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8", "DNT": "1", "Connection": "close", "Upgrade-Insecure-Requests": "1"}


##!  vreate a html file from Website Amazon


# Send a GET request to the URL
response = requests.get(url, headers=headers)


# Check if the request was successful
if response.status_code == 200:
    # Write the webpage content to a file
    with open(filename, 'w', encoding='utf-8') as file:
        file.write(response.text)
    print(f"The webpage content was saved as {filename} successfully.")
else:
    print(f"Failed to retrieve the webpage. {response.status_code}")

########! end file crate start Scraping

# Read the HTML file
with open(filename, 'r', encoding='utf-8') as file:
    html = file.read()

# Parse the HTML using BeautifulSoup

soup = BeautifulSoup(html, 'html.parser')


# Find all divs with class="sg-col-inner"
divs = soup.find_all(
    'div', class_='a-section a-spacing-small a-spacing-top-small')
# divs = soup.find_all('div', class_='sg-col-inner')

# Create an Excel workbook and select the active sheet
workbook = openpyxl.Workbook()
sheet = workbook.active

# Write headers
sheet['A1'] = 'Product Name'
sheet['B1'] = 'Product Price'

# Initialize row counter
row = 2

# Iterate over the divs
for div in divs:
    # Find the span with class="a-size-medium a-color-base a-text-normal" for product name
    product_name_span = div.find(
        'span', class_='a-size-medium a-color-base a-text-normal')
    product_name = product_name_span.text.strip() if product_name_span else ' '

    # Find the span with class="a-price-whole" for product price
    product_price_span = div.find('span', class_='a-price-whole')
    product_price = product_price_span.text.strip() if product_price_span else '0'

    product_price_span2 = div.find('span', class_='a-price-fraction')
    product_price2 = product_price_span2.text.strip() if product_price_span2 else '0'

    # Write product name and price to Excel
    sheet.cell(row=row, column=1).value = product_name
    sheet.cell(row=row, column=2).value = product_price + product_price2

    # Increment row counter
    row += 1

# Save the Excel file
workbook.save(f'Product_Data{x}.xlsx')
print(f"The Product_Data{x}.xlsx saved successfully.")
##!
# # Check if the file already exists
# filename = f"Product_Data_{today}.xlsx"
# counter = 1
# while os.path.exists(filename):
#     # If the file already exists, increment the counter and generate a new file name
#     counter += 1
#     filename = f"Product_Data_{today}_{counter}.xlsx"

# # Rest of the code to retrieve and parse the HTML

# # Save the Excel file
# workbook.save(filename)
# print(f"The product data was saved as {filename} successfully.")
